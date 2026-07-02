from __future__ import annotations

import csv
import io
from collections import Counter, defaultdict
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROCESSED_FIELDS = [
    "Organization", "Storefront", "Batch Number", "Order ID", "External Order #",
    "Package Name", "Package Sequence", "Order Status", "Tracking Number",
    "Tracking Status", "Fulfillment Status", "Items Ordered (Total)", "Items Fulfilled (This Pkg)", "SKUs in Package",
    "Fulfillment Fee", "Packaging Fee", "Dimensions (LxWxH)",
    "Order Date", "Fulfillment Date", "Notes",
]

INVOICE_FIELDS = [
    "Invoice ID", "InvoiceItem ID", "Invoice Number", "Invoice Desc", "Created",
    "Type", "Description", "Batch", "Amount", "Qty", "Batch Number",
    "Order Number", "Num Items",
]

OUTPUT_FIELDS = [
    "Reconciliation Status", "External Order #", "Processed Batch Number",
    "Invoice Order Number", "Invoice Batch Number", "Invoice Number", "Invoice Amount",
    "Fulfillment Costs", "Packaging Fees", "Payment Service Fees", "Mapped Invoice Total",
    "Invoice Qty", "Invoice Num Items", "Processed Order ID", "Storefront", "Package Count",
    "Package Names", "Package Dimensions", "Tracking Numbers", "Items Ordered", "Items Fulfilled",
    "SKUs in Package", "Processed Fulfillment Fee", "Processed Packaging Fee",
    "Order Date", "Fulfillment Date",
    "Order Status", "Tracking Status", "Fulfillment Status", "Notes", "Match Note",
]

NA_BATCH_FEE_FIELDS = [
    "First Pick Fee", "Additional Picks", "Additional Pick Fee",
    "Calculated Fulfillment Total", "Calculated Packaging Fee",
    "Packaging Pricing Status", "Estimated Fulfillment + Packaging",
]
NA_BATCH_OUTPUT_FIELDS = OUTPUT_FIELDS.copy()
_na_fee_index = NA_BATCH_OUTPUT_FIELDS.index("Items Fulfilled") + 1
NA_BATCH_OUTPUT_FIELDS[_na_fee_index:_na_fee_index] = NA_BATCH_FEE_FIELDS
ORDER_INVOICE_FEE_TYPES = {"fulfillmentfees", "supplies"}
PAYMENT_SERVICE_FEE_TYPE = "paymentservicefees"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def match_id(value: Any) -> str:
    value_text = text(value)
    if re.fullmatch(r"\d+\.0+", value_text):
        value_text = value_text.split(".", 1)[0]
    if re.fullmatch(r"\d+", value_text):
        return value_text.lstrip("0") or "0"
    return re.sub(r"\s+", "", value_text).upper()


def is_na_batch(value: Any) -> bool:
    value_text = re.sub(r"[^A-Z0-9]+", "", text(value).upper())
    return value_text in {"", "NA", "NOTAVAILABLE", "NOTAPPLICABLE"}


def parse_source_date(value: Any) -> date | None:
    value_text = text(value)
    if not value_text:
        return None
    for candidate in (value_text, value_text[:10]):
        for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(candidate, date_format).date()
            except ValueError:
                continue
    return None


def decimal_value(value: Any) -> Decimal:
    value_text = text(value).replace(",", "").replace("$", "")
    if not value_text:
        return Decimal("0")
    try:
        return Decimal(value_text)
    except InvalidOperation:
        return Decimal("0")


def display_number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral() else float(value)


def sum_field(rows: list[dict[str, str]], field: str) -> Decimal:
    return sum((decimal_value(row.get(field, "")) for row in rows), Decimal("0"))


def invoice_type_amount(rows: list[dict[str, str]], fee_type: str) -> Decimal:
    return sum(
        (
            decimal_value(row.get("Amount", ""))
            for row in rows
            if header_key(row.get("Type", "")) == fee_type
        ),
        Decimal("0"),
    )


def invoice_fee_values(rows: list[dict[str, str]]) -> dict[str, int | float]:
    fulfillment = invoice_type_amount(rows, "fulfillmentfees")
    packaging = invoice_type_amount(rows, "supplies")
    payment_service = invoice_type_amount(rows, PAYMENT_SERVICE_FEE_TYPE)
    mapped_total = fulfillment + packaging + payment_service
    return {
        "Invoice Amount": display_number(mapped_total),
        "Fulfillment Costs": display_number(fulfillment),
        "Packaging Fees": display_number(packaging),
        "Payment Service Fees": display_number(payment_service),
        "Mapped Invoice Total": display_number(mapped_total),
    }


def normalize_dimensions(value: Any) -> str:
    numbers = re.findall(r"\d+(?:\.\d+)?", text(value))
    if len(numbers) < 3:
        return ""
    normalized_numbers = []
    for number in numbers[:3]:
        decimal_number = Decimal(number)
        normalized_numbers.append(format(decimal_number.normalize(), "f"))
    return "x".join(normalized_numbers)


def build_box_pricing(
    processed_rows: list[dict[str, str]],
    invoice_rows: list[dict[str, str]],
) -> dict[str, dict[str, Any]]:
    processed_by_key: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in processed_rows:
        processed_by_key[(
            match_id(row.get("External Order #", "")),
            match_id(row.get("Batch Number", "")),
        )].append(row)
    samples: dict[str, list[Decimal]] = defaultdict(list)
    for invoice_row in invoice_rows:
        if header_key(invoice_row.get("Type", "")) != "supplies":
            continue
        source_rows = processed_by_key.get((
            match_id(invoice_row.get("Order Number", "")),
            match_id(invoice_row.get("Batch Number", "")),
        ), [])
        if len(source_rows) != 1:
            continue
        dimension = normalize_dimensions(
            source_rows[0].get("Dimensions (LxWxH)", "")
            or source_rows[0].get("Package Name", "")
        )
        amount = decimal_value(invoice_row.get("Amount", ""))
        if dimension and amount > 0:
            samples[dimension].append(amount)
    lookup: dict[str, dict[str, Any]] = {}
    for dimension, prices in samples.items():
        counts = Counter(prices)
        price, winning_samples = counts.most_common(1)[0]
        lookup[dimension] = {
            "price": display_number(price),
            "samples": len(prices),
            "matching_samples": winning_samples,
            "confidence": winning_samples / len(prices),
        }
    return lookup


def join_field(rows: list[dict[str, str]], field: str) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = text(row.get(field, ""))
        if value and value not in seen:
            seen.add(value)
            values.append(value)
    return ", ".join(values)


def decode_csv(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("latin-1")


def read_csv_rows(filename: str, data: bytes, required_fields: list[str]) -> list[dict[str, str]]:
    if Path(filename).suffix.lower() != ".csv":
        raise ValueError(f"{filename}: upload a CSV file.")
    reader = csv.DictReader(io.StringIO(decode_csv(data), newline=""))
    if not reader.fieldnames:
        raise ValueError(f"{filename}: no header row was found.")
    actual_by_key = {header_key(field): field for field in reader.fieldnames}
    missing = [field for field in required_fields if header_key(field) not in actual_by_key]
    if missing:
        raise ValueError(f"{filename}: missing required columns: {', '.join(missing)}.")
    rows: list[dict[str, str]] = []
    for source in reader:
        row = {field: text(source.get(actual_by_key[header_key(field)], "")) for field in required_fields}
        if any(row.values()):
            rows.append(row)
    return rows


def aggregate_processed(
    rows: list[dict[str, str]],
    status: str,
    invoice_rows: list[dict[str, str]] | None = None,
    invoice_batches: str = "",
    note: str = "",
) -> dict[str, Any]:
    invoices = invoice_rows or []
    items_ordered = max((decimal_value(row.get("Items Ordered (Total)", "")) for row in rows), default=Decimal("0"))
    return {
        "Reconciliation Status": status,
        "External Order #": join_field(rows, "External Order #"),
        "Processed Batch Number": join_field(rows, "Batch Number"),
        "Invoice Order Number": join_field(invoices, "Order Number"),
        "Invoice Batch Number": join_field(invoices, "Batch Number") or invoice_batches,
        "Invoice Number": join_field(invoices, "Invoice Number"),
        **invoice_fee_values(invoices),
        "Invoice Qty": display_number(sum_field(invoices, "Qty")),
        "Invoice Num Items": display_number(sum_field(invoices, "Num Items")),
        "Processed Order ID": join_field(rows, "Order ID"),
        "Storefront": join_field(rows, "Storefront"),
        "Package Count": len(rows),
        "Package Names": join_field(rows, "Package Name"),
        "Package Dimensions": join_field(rows, "Dimensions (LxWxH)"),
        "Tracking Numbers": join_field(rows, "Tracking Number"),
        "Items Ordered": display_number(items_ordered),
        "Items Fulfilled": display_number(sum_field(rows, "Items Fulfilled (This Pkg)")),
        "SKUs in Package": join_field(rows, "SKUs in Package"),
        "Processed Fulfillment Fee": display_number(sum_field(rows, "Fulfillment Fee")),
        "Processed Packaging Fee": display_number(sum_field(rows, "Packaging Fee")),
        "Order Date": join_field(rows, "Order Date"),
        "Fulfillment Date": join_field(rows, "Fulfillment Date"),
        "Order Status": join_field(rows, "Order Status"),
        "Tracking Status": join_field(rows, "Tracking Status"),
        "Fulfillment Status": join_field(rows, "Fulfillment Status"),
        "Notes": join_field(rows, "Notes"),
        "Match Note": note,
    }


def invoice_unmatched_row(
    invoice_rows: list[dict[str, str]],
    processed_rows: list[dict[str, str]],
    processed_batches: str,
    note: str,
) -> dict[str, Any]:
    return {
        "Reconciliation Status": "Invoice Unmatched",
        "External Order #": join_field(invoice_rows, "Order Number"),
        "Processed Batch Number": processed_batches,
        "Invoice Order Number": join_field(invoice_rows, "Order Number"),
        "Invoice Batch Number": join_field(invoice_rows, "Batch Number"),
        "Invoice Number": join_field(invoice_rows, "Invoice Number"),
        **invoice_fee_values(invoice_rows),
        "Invoice Qty": display_number(sum_field(invoice_rows, "Qty")),
        "Invoice Num Items": display_number(sum_field(invoice_rows, "Num Items")),
        "Processed Order ID": "",
        "Storefront": join_field(invoice_rows, "Batch"),
        "Package Count": "",
        "Package Names": "",
        "Package Dimensions": "",
        "Tracking Numbers": "",
        "Items Ordered": "",
        "Items Fulfilled": "",
        "SKUs in Package": "",
        "Processed Fulfillment Fee": "",
        "Processed Packaging Fee": display_number(sum_field(processed_rows, "Packaging Fee")),
        "Order Date": "",
        "Fulfillment Date": "",
        "Order Status": join_field(processed_rows, "Order Status"),
        "Tracking Status": join_field(processed_rows, "Tracking Status"),
        "Fulfillment Status": join_field(processed_rows, "Fulfillment Status"),
        "Notes": join_field(processed_rows, "Notes"),
        "Match Note": note,
    }


def apply_na_batch_fees(row: dict[str, Any]) -> None:
    fulfilled_items = max(decimal_value(row.get("Items Fulfilled", "")), Decimal("0"))
    first_pick_fee = Decimal("2.25") if fulfilled_items > 0 else Decimal("0")
    additional_picks = max(fulfilled_items - Decimal("1"), Decimal("0"))
    additional_pick_fee = additional_picks * Decimal("0.50")
    row["First Pick Fee"] = display_number(first_pick_fee)
    row["Additional Picks"] = display_number(additional_picks)
    row["Additional Pick Fee"] = display_number(additional_pick_fee)
    row["Calculated Fulfillment Total"] = display_number(first_pick_fee + additional_pick_fee)


def apply_na_batch_packaging(
    row: dict[str, Any],
    source_rows: list[dict[str, str]],
    box_pricing: dict[str, dict[str, Any]],
) -> None:
    calculated_fee = Decimal("0")
    statuses: list[str] = []
    unresolved = False
    for source in source_rows:
        source_fee = decimal_value(source.get("Packaging Fee", ""))
        if source_fee > 0:
            calculated_fee += source_fee
            if "Column AA Packaging Fee" not in statuses:
                statuses.append("Column AA Packaging Fee")
            continue
        package_name = text(source.get("Package Name", ""))
        dimension = normalize_dimensions(
            source.get("Dimensions (LxWxH)", "") or package_name
        )
        if "nopackagingselected" in header_key(package_name):
            if "No packaging selected" not in statuses:
                statuses.append("No packaging selected")
            continue
        pricing = box_pricing.get(dimension)
        if pricing:
            calculated_fee += decimal_value(pricing["price"])
            status = f"Invoice dimension lookup: {dimension} (n={pricing['samples']})"
            if status not in statuses:
                statuses.append(status)
        else:
            unresolved = True
            status = f"No pricing found: {dimension or package_name or 'blank package'}"
            if status not in statuses:
                statuses.append(status)
    row["Calculated Packaging Fee"] = display_number(calculated_fee)
    row["Packaging Pricing Status"] = "; ".join(statuses)
    if unresolved:
        row["Estimated Fulfillment + Packaging"] = ""
    else:
        total = decimal_value(row.get("Calculated Fulfillment Total", "")) + calculated_fee
        row["Estimated Fulfillment + Packaging"] = display_number(total)


def add_box_pricing_sheet(
    workbook: Workbook,
    box_pricing: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Box Pricing")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    headers = ["Dimensions", "Price", "Clean Samples", "Matching Samples", "Confidence", "Source"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor="10233F")
        cell.font = Font(color="FFFFFF", bold=True)
    for row_index, (dimension, pricing) in enumerate(
        sorted(box_pricing.items(), key=lambda item: (-item[1]["samples"], item[0])),
        start=2,
    ):
        values = [
            dimension, pricing["price"], pricing["samples"],
            pricing["matching_samples"], pricing["confidence"],
            "Invoice supplies joined to a single-package processed order",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.fill = PatternFill("solid", fgColor="E8F7EF")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=2).number_format = '"$"#,##0.00'
        sheet.cell(row=row_index, column=5).number_format = "0.0%"
    for column, width in {"A": 18, "B": 12, "C": 15, "D": 17, "E": 13, "F": 58}.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = f"A1:F{max(1, len(box_pricing) + 1)}"


def add_data_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    fill_color: str,
    fields: list[str] | None = None,
) -> None:
    fields = fields or OUTPUT_FIELDS
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="10233F")
    body_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E3F0"))
    for column_index, field in enumerate(fields, start=1):
        cell = sheet.cell(row=1, column=column_index, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fields, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row.get(field, ""))
            cell.fill = body_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if field in {"Invoice Amount", "Fulfillment Costs", "Packaging Fees", "Payment Service Fees", "Mapped Invoice Total", "Processed Fulfillment Fee", "Processed Packaging Fee", "First Pick Fee", "Additional Pick Fee", "Calculated Fulfillment Total", "Calculated Packaging Fee", "Estimated Fulfillment + Packaging"} and isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'
            elif field in {"Package Count", "Items Ordered", "Items Fulfilled", "Invoice Qty", "Invoice Num Items", "Additional Picks"} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
    last_row = max(1, len(rows) + 1)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{last_row}"
    widths = {
        "Reconciliation Status": 22, "External Order #": 19, "Processed Batch Number": 18,
        "Invoice Order Number": 19, "Invoice Batch Number": 18, "Invoice Number": 24,
        "Invoice Amount": 15, "Fulfillment Costs": 17, "Packaging Fees": 16,
        "Payment Service Fees": 20, "Mapped Invoice Total": 19,
        "Invoice Qty": 12, "Invoice Num Items": 16,
        "Processed Order ID": 18, "Storefront": 26, "Package Count": 14,
        "Package Names": 26, "Package Dimensions": 20, "Tracking Numbers": 28, "Items Ordered": 14,
        "Items Fulfilled": 15, "SKUs in Package": 42, "Processed Fulfillment Fee": 20,
        "Processed Packaging Fee": 20, "Calculated Packaging Fee": 20,
        "Packaging Pricing Status": 46, "Estimated Fulfillment + Packaging": 28,
        "Order Date": 14, "Fulfillment Date": 16, "Order Status": 16,
        "Tracking Status": 18, "Fulfillment Status": 18, "Notes": 42,
        "First Pick Fee": 15, "Additional Picks": 16, "Additional Pick Fee": 18,
        "Calculated Fulfillment Total": 23, "Match Note": 42,
    }
    for column_index, field in enumerate(fields, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = widths.get(field, 16)
    sheet.row_dimensions[1].height = 34


def write_workbook(
    output_dir: Path,
    matched: list[dict[str, Any]],
    batch_mismatches: list[dict[str, Any]],
    processed_unmatched: list[dict[str, Any]],
    na_batch: list[dict[str, Any]],
    invoice_unmatched: list[dict[str, Any]],
    processed_filename: str,
    invoice_filename: str,
    completed_source_rows: int,
    cancelled_ignored: int,
    other_status_ignored: int,
    invoice_fee_summary: dict[str, Any],
    box_pricing: dict[str, dict[str, Any]],
    start_date: date,
    end_date: date,
    processed_date_excluded: int,
    invoice_date_excluded: int,
) -> Path:
    na_items_fulfilled = sum((decimal_value(row.get("Items Fulfilled", "")) for row in na_batch), Decimal("0"))
    na_first_pick_fees = sum((decimal_value(row.get("First Pick Fee", "")) for row in na_batch), Decimal("0"))
    na_additional_picks = sum((decimal_value(row.get("Additional Picks", "")) for row in na_batch), Decimal("0"))
    na_additional_pick_fees = sum((decimal_value(row.get("Additional Pick Fee", "")) for row in na_batch), Decimal("0"))
    na_calculated_total = sum((decimal_value(row.get("Calculated Fulfillment Total", "")) for row in na_batch), Decimal("0"))
    na_packaging_total = sum((decimal_value(row.get("Calculated Packaging Fee", "")) for row in na_batch), Decimal("0"))
    na_estimated_total = sum((decimal_value(row.get("Estimated Fulfillment + Packaging", "")) for row in na_batch), Decimal("0"))
    na_inferred_count = sum(1 for row in na_batch if "Invoice dimension lookup" in text(row.get("Packaging Pricing Status", "")))
    na_no_package_count = sum(1 for row in na_batch if "No packaging selected" in text(row.get("Packaging Pricing Status", "")))
    na_missing_pricing_count = sum(1 for row in na_batch if "No pricing found" in text(row.get("Packaging Pricing Status", "")))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "Christy WH Invoice Reconciliation"
    summary["A1"].fill = PatternFill("solid", fgColor="10233F")
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 30
    summary_rows = [
        ("Generated", date.today().isoformat(), "", ""),
        ("Selected Date Range", f"{start_date.isoformat()} through {end_date.isoformat()}", "Processed Fulfillment Date (Order Date fallback) / Invoice Created", ""),
        ("Processed Orders Source", processed_filename, "", ""),
        ("Invoice Export Source", invoice_filename, "", ""),
        ("", "", "", ""),
        ("Category", "Order/Batch Groups", "Meaning", "Review Priority"),
        ("Matched", len(matched), "Order number and batch number both match.", "Ready to invoice"),
        ("Batch Mismatches", len(batch_mismatches), "Order exists in both exports but batch numbers differ.", "Review"),
        ("Processed Unmatched", len(processed_unmatched), "Completed processed order is not present in the invoice export.", "Review"),
        ("N/A Batch", len(na_batch), "Completed processed order has N/A or blank batch.", "Check outside-platform labels"),
        ("Invoice Unmatched", len(invoice_unmatched), "Invoice order/batch has no matching completed processed order/batch.", "Review"),
        ("", "", "", ""),
        ("Invoice Fee Summary", "", "Mapped from the invoice Amount column by Type.", ""),
        ("Fulfillment Costs", invoice_fee_summary["fulfillment_costs"], "Type = fulfillment_fees", ""),
        ("Packaging Fees", invoice_fee_summary["packaging_fees"], "Type = supplies", ""),
        ("Payment Service Fees", invoice_fee_summary["payment_service_fees"], "Type = payment_service_fees; usually one invoice-level row.", ""),
        ("Mapped Invoice Fee Total", invoice_fee_summary["mapped_total"], "Fulfillment + packaging + payment service fees.", ""),
        ("Other Invoice Fees Excluded", invoice_fee_summary["other_fees"], f"{invoice_fee_summary['other_fee_rows']} returns/return-label or other rows excluded from order matching.", ""),
        ("", "", "", ""),
        ("N/A Batch Fee Summary", "", "$2.25 first pick + $0.50 per additional fulfilled item", ""),
        ("N/A Fulfilled Items", display_number(na_items_fulfilled), "Total fulfilled items across grouped N/A orders.", ""),
        ("N/A First Pick Fees", display_number(na_first_pick_fees), "One $2.25 first-pick fee per order with fulfilled items.", ""),
        ("N/A Additional Picks", display_number(na_additional_picks), "Fulfilled items after the first item.", ""),
        ("N/A Additional Pick Fees", display_number(na_additional_pick_fees), "$0.50 per additional pick.", ""),
        ("N/A Calculated Fulfillment Total", display_number(na_calculated_total), "First-pick fees plus additional-pick fees.", ""),
        ("N/A Calculated Packaging Fees", display_number(na_packaging_total), "Column AA fees plus invoice-mined dimension pricing where needed.", ""),
        ("N/A Estimated Fulfillment + Packaging", display_number(na_estimated_total), "Fulfillment plus known packaging fees; unresolved packaging orders are excluded.", ""),
        ("N/A Packaging Inferred", na_inferred_count, "Orders priced from the invoice-mined dimension lookup.", ""),
        ("N/A No Packaging Selected", na_no_package_count, "Orders explicitly marked No Packaging selected.", ""),
        ("N/A Missing Packaging Pricing", na_missing_pricing_count, "Orders with packaging but no usable price.", "Review"),
        ("Box Dimensions Mined", len(box_pricing), "See the Box Pricing sheet for rates and sample counts.", ""),
        ("", "", "", ""),
        ("Completed Processed Source Rows", completed_source_rows, "Includes multiple package rows before grouping.", ""),
        ("Cancelled Rows Ignored", cancelled_ignored, "Excluded from all reconciliation tabs.", ""),
        ("Other Status Rows Ignored", other_status_ignored, "Only Complete or Completed statuses are reconciled.", ""),
        ("Processed Rows Outside Date Range", processed_date_excluded, "Fulfillment Date, or fallback Order Date, is outside the selected range or missing.", ""),
        ("Invoice Rows Outside Date Range", invoice_date_excluded, "Created date is outside the selected range or missing.", ""),
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="10233F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for row_index, fill_color in {8: "E8F7EF", 9: "FFF2CC", 10: "FDE9E7", 11: "FFF4D6", 12: "FDE9E7"}.items():
        for cell in summary[row_index]:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    section_labels = {"Invoice Fee Summary", "N/A Batch Fee Summary"}
    total_labels = {"Mapped Invoice Fee Total", "N/A Calculated Fulfillment Total", "N/A Estimated Fulfillment + Packaging"}
    currency_labels = {
        "Fulfillment Costs", "Packaging Fees", "Payment Service Fees",
        "Mapped Invoice Fee Total", "Other Invoice Fees Excluded",
        "N/A First Pick Fees", "N/A Additional Pick Fees",
        "N/A Calculated Fulfillment Total", "N/A Calculated Packaging Fees",
        "N/A Estimated Fulfillment + Packaging",
    }
    for row_index in range(1, summary.max_row + 1):
        label = summary.cell(row=row_index, column=1).value
        if label in section_labels:
            for cell in summary[row_index]:
                cell.fill = PatternFill("solid", fgColor="10233F")
                cell.font = Font(color="FFFFFF", bold=True)
        if label in currency_labels:
            summary.cell(row=row_index, column=2).number_format = '"$"#,##0.00'
        if label in total_labels:
            for cell in summary[row_index]:
                cell.fill = PatternFill("solid", fgColor="E8F7EF")
                cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 64
    summary.column_dimensions["D"].width = 27
    summary.freeze_panes = "A7"

    add_data_sheet(workbook, "Matched", matched, "E8F7EF")
    add_data_sheet(workbook, "Batch Mismatches", batch_mismatches, "FFF2CC")
    add_data_sheet(workbook, "Processed Unmatched", processed_unmatched, "FDE9E7")
    add_data_sheet(workbook, "N-A Batch", na_batch, "FFF4D6", NA_BATCH_OUTPUT_FIELDS)
    add_data_sheet(workbook, "Invoice Unmatched", invoice_unmatched, "FDE9E7")
    add_box_pricing_sheet(workbook, box_pricing)
    output_path = output_dir / f"Christy_WH_Invoice_Reconciliation_{uuid.uuid4().hex[:8]}.xlsx"
    workbook.save(output_path)
    return output_path


def build_christy_invoice_reconciliation(
    processed_filename: str,
    processed_data: bytes,
    invoice_filename: str,
    invoice_data: bytes,
    output_dir: Path,
    start_date_value: str,
    end_date_value: str,
) -> dict[str, Any]:
    processed_rows = read_csv_rows(processed_filename, processed_data, PROCESSED_FIELDS)
    invoice_rows = read_csv_rows(invoice_filename, invoice_data, INVOICE_FIELDS)
    start_date = parse_source_date(start_date_value)
    end_date = parse_source_date(end_date_value)
    if start_date is None or end_date is None:
        raise ValueError("Select a valid start date and end date.")
    if start_date > end_date:
        raise ValueError("Start date must be on or before end date.")

    processed_rows_in_range = [
        row for row in processed_rows
        if (
            row_date := (
                parse_source_date(row.get("Fulfillment Date", ""))
                or parse_source_date(row.get("Order Date", ""))
            )
        ) is not None
        and start_date <= row_date <= end_date
    ]
    invoice_rows_in_range = [
        row for row in invoice_rows
        if (row_date := parse_source_date(row.get("Created", ""))) is not None
        and start_date <= row_date <= end_date
    ]
    processed_date_excluded = len(processed_rows) - len(processed_rows_in_range)
    invoice_date_excluded = len(invoice_rows) - len(invoice_rows_in_range)

    completed_rows: list[dict[str, str]] = []
    cancelled_ignored = 0
    other_status_ignored = 0
    for row in processed_rows_in_range:
        status = header_key(row.get("Order Status", ""))
        if status in {"complete", "completed"}:
            completed_rows.append(row)
        elif status in {"cancelled", "canceled"}:
            cancelled_ignored += 1
        else:
            other_status_ignored += 1

    order_fee_rows = [
        row for row in invoice_rows_in_range
        if header_key(row.get("Type", "")) in ORDER_INVOICE_FEE_TYPES
    ]
    payment_service_rows = [
        row for row in invoice_rows_in_range
        if header_key(row.get("Type", "")) == PAYMENT_SERVICE_FEE_TYPE
    ]
    other_fee_rows = [
        row for row in invoice_rows_in_range
        if header_key(row.get("Type", "")) not in ORDER_INVOICE_FEE_TYPES | {PAYMENT_SERVICE_FEE_TYPE}
    ]
    fulfillment_total = invoice_type_amount(order_fee_rows, "fulfillmentfees")
    packaging_total = invoice_type_amount(order_fee_rows, "supplies")
    payment_service_total = invoice_type_amount(payment_service_rows, PAYMENT_SERVICE_FEE_TYPE)
    other_fee_total = sum_field(other_fee_rows, "Amount")
    invoice_fee_summary = {
        "fulfillment_costs": display_number(fulfillment_total),
        "packaging_fees": display_number(packaging_total),
        "payment_service_fees": display_number(payment_service_total),
        "mapped_total": display_number(fulfillment_total + packaging_total + payment_service_total),
        "other_fees": display_number(other_fee_total),
        "other_fee_rows": len(other_fee_rows),
    }

    box_pricing = build_box_pricing(processed_rows, order_fee_rows)

    invoice_by_key: dict[tuple[str, str], list[dict[str, str]]] = {}
    invoice_by_order: dict[str, list[dict[str, str]]] = {}
    for row in order_fee_rows:
        order_key = match_id(row.get("Order Number", ""))
        batch_key = match_id(row.get("Batch Number", ""))
        invoice_by_key.setdefault((order_key, batch_key), []).append(row)
        invoice_by_order.setdefault(order_key, []).append(row)

    valid_groups: dict[tuple[str, str], list[dict[str, str]]] = {}
    na_groups: dict[str, list[dict[str, str]]] = {}
    for row in completed_rows:
        order_key = match_id(row.get("External Order #", ""))
        fallback = match_id(row.get("Order ID", "")) or uuid.uuid4().hex
        if is_na_batch(row.get("Batch Number", "")):
            na_groups.setdefault(order_key or f"missing-{fallback}", []).append(row)
        else:
            batch_key = match_id(row.get("Batch Number", ""))
            valid_groups.setdefault((order_key or f"missing-{fallback}", batch_key), []).append(row)

    matched: list[dict[str, Any]] = []
    batch_mismatches: list[dict[str, Any]] = []
    processed_unmatched: list[dict[str, Any]] = []
    for (order_key, batch_key), rows in valid_groups.items():
        exact_invoices = invoice_by_key.get((order_key, batch_key), [])
        if exact_invoices:
            matched.append(aggregate_processed(rows, "Matched", exact_invoices, note="Order number and batch number match the invoice export."))
        elif invoice_by_order.get(order_key):
            order_invoices = invoice_by_order[order_key]
            invoice_batches = join_field(order_invoices, "Batch Number")
            batch_mismatches.append(aggregate_processed(rows, "Batch Mismatch", invoice_batches=invoice_batches, note=f"Order number exists in the invoice export under batch {invoice_batches or 'blank'}."))
        else:
            processed_unmatched.append(aggregate_processed(rows, "Processed Unmatched", note="Completed processed order was not found in the invoice export."))

    na_batch: list[dict[str, Any]] = []
    for source_rows in na_groups.values():
        row = aggregate_processed(
            source_rows,
            "N/A Batch - Review",
            note="Completed order has N/A or blank batch; verify whether its label was processed outside the platform.",
        )
        apply_na_batch_fees(row)
        apply_na_batch_packaging(row, source_rows, box_pricing)
        na_batch.append(row)

    processed_by_key = set(valid_groups)
    processed_by_order: dict[str, set[str]] = {}
    for order_key, batch_key in processed_by_key:
        processed_by_order.setdefault(order_key, set()).add(batch_key)
    all_processed_by_order: dict[str, list[dict[str, str]]] = {}
    for row in processed_rows_in_range:
        all_processed_by_order.setdefault(match_id(row.get("External Order #", "")), []).append(row)
    invoice_unmatched: list[dict[str, Any]] = []
    for (order_key, batch_key), fee_rows in invoice_by_key.items():
        if (order_key, batch_key) in processed_by_key:
            continue
        batches = ", ".join(sorted(processed_by_order.get(order_key, set())))
        note = "Order exists in completed processed orders, but the batch number differs." if batches else "No matching completed processed order was found."
        invoice_unmatched.append(
            invoice_unmatched_row(
                fee_rows,
                all_processed_by_order.get(order_key, []),
                batches,
                note,
            )
        )

    sort_key = lambda row: (text(row.get("Processed Batch Number", "")), text(row.get("External Order #", "")))
    for rows in [matched, batch_mismatches, processed_unmatched, na_batch, invoice_unmatched]:
        rows.sort(key=sort_key)

    output_path = write_workbook(
        output_dir, matched, batch_mismatches, processed_unmatched, na_batch,
        invoice_unmatched, processed_filename, invoice_filename, len(completed_rows),
        cancelled_ignored, other_status_ignored, invoice_fee_summary, box_pricing, start_date, end_date,
        processed_date_excluded, invoice_date_excluded,
    )
    review_preview = (batch_mismatches + processed_unmatched + na_batch + invoice_unmatched)[:100]
    return {
        "download_url": f"/download/{output_path.name}",
        "filename": output_path.name,
        "counts": {
            "matched": len(matched),
            "batch_mismatches": len(batch_mismatches),
            "processed_unmatched": len(processed_unmatched),
            "na_batch": len(na_batch),
            "invoice_unmatched": len(invoice_unmatched),
            "cancelled_ignored": cancelled_ignored,
            "other_status_ignored": other_status_ignored,
            "processed_date_excluded": processed_date_excluded,
            "invoice_date_excluded": invoice_date_excluded,
        },
        "date_range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "invoice_fee_summary": invoice_fee_summary,
        "na_packaging_summary": {
            "packaging_fees": display_number(sum((decimal_value(row.get("Calculated Packaging Fee", "")) for row in na_batch), Decimal("0"))),
            "estimated_total": display_number(sum((decimal_value(row.get("Estimated Fulfillment + Packaging", "")) for row in na_batch), Decimal("0"))),
            "inferred": sum(1 for row in na_batch if "Invoice dimension lookup" in text(row.get("Packaging Pricing Status", ""))),
            "no_packaging_selected": sum(1 for row in na_batch if "No packaging selected" in text(row.get("Packaging Pricing Status", ""))),
            "missing_pricing": sum(1 for row in na_batch if "No pricing found" in text(row.get("Packaging Pricing Status", ""))),
            "dimensions_mined": len(box_pricing),
        },
        "box_pricing_preview": [
            {"Dimensions": dimension, **pricing}
            for dimension, pricing in sorted(box_pricing.items(), key=lambda item: (-item[1]["samples"], item[0]))
        ],
        "matched_preview": matched[:50],
        "review_preview": review_preview,
    }
